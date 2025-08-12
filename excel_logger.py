# excel_logger.py

import os
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# Importa desde config y data_manager
from config import EXCEL_LOG_FILE_FULL_PATH, LOGS_PATH
import data_manager

def inicializar_excel_log():
    """Crea el archivo Excel y la carpeta de logs si no existen."""
    if not os.path.exists(LOGS_PATH):
        try:
            os.makedirs(LOGS_PATH, exist_ok=True)
        except OSError as e:
            print(f"ADVERTENCIA: No se pudo crear el directorio de logs '{LOGS_PATH}': {e}")

    if not os.path.exists(EXCEL_LOG_FILE_FULL_PATH):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Registro de Actividad"
            headers = ["Timestamp", "Usuario", "Rol", "Accion", "Detalles Adicionales", "Duracion Sesion (min)"]
            sheet.append(headers)
            for col_num, header_title in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                column_letter = get_column_letter(col_num)
                if header_title == "Timestamp": sheet.column_dimensions[column_letter].width = 20
                elif header_title == "Detalles Adicionales": sheet.column_dimensions[column_letter].width = 45
                elif header_title == "Duracion Sesion (min)": sheet.column_dimensions[column_letter].width = 22
                else: sheet.column_dimensions[column_letter].width = 25
            workbook.save(EXCEL_LOG_FILE_FULL_PATH)
        except Exception as e:
            print(f"ADVERTENCIA: No se pudo inicializar el archivo de log Excel '{EXCEL_LOG_FILE_FULL_PATH}': {e}")

def registrar_accion_excel(accion, detalles="", duracion_min=None):
    """Registra una acción en el archivo Excel."""
    if not os.path.exists(LOGS_PATH):
        inicializar_excel_log()

    try:
        # Cargar o crear el libro de trabajo
        if os.path.exists(EXCEL_LOG_FILE_FULL_PATH):
            workbook = openpyxl.load_workbook(EXCEL_LOG_FILE_FULL_PATH)
            sheet = workbook.active
        else:
            print(f"ADVERTENCIA: Archivo de log '{EXCEL_LOG_FILE_FULL_PATH}' no encontrado. Creando uno nuevo para esta sesión.")
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Registro de Actividad"
            headers = ["Timestamp", "Usuario", "Rol", "Accion", "Detalles Adicionales", "Duracion Sesion (min)"]
            sheet.append(headers)
    
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        nombre_usuario_log = data_manager.usuario_actual.get("nombre", "N/A")
        rol_usuario_log = data_manager.usuario_actual.get("rol", "N/A")
        
        log_entry = [timestamp, nombre_usuario_log, rol_usuario_log, accion, detalles,
                    f"{duracion_min:.2f}" if duracion_min is not None else ""]
        sheet.append(log_entry)
        workbook.save(EXCEL_LOG_FILE_FULL_PATH)

    except Exception as e:
        print(f"ERROR al registrar acción en Excel: {e}. Acción: '{accion}', Detalles: '{detalles}'")
